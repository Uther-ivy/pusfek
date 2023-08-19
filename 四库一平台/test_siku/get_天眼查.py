import logging
import random
import time
import traceback
import urllib.parse

import requests

import xlwt
from lxml import etree
from openpyxl.reader.excel import load_workbook

from spider import ip_proxys


class get_company(object):

    def __init__(self):
        self._session=requests.session()

        self.proxys = {
            'http': '',
            'https': ''
        }
        self.headers= {
        'cookie':'jsid=SEO-BAIDU-ALL-SY-000001; TYCID=4e774e20d4ee11edb16713c865b4927b; ssuid=9093622780; _ga=GA1.2.160854741.1680835765; HWWAFSESID=c2f1549cea3d805c7b9; HWWAFSESTIME=1681725716648; csrfToken=YVtFSs_WNW9bDYowcvH8SjtO; bdHomeCount=3; Hm_lvt_e92c8d65d92d534b0fc290df538b4758=1680835551,1681725719; bannerFlag=true; searchSessionId=1681725731.76420050; sensorsdata2015jssdkcross=%7B%22distinct_id%22%3A%2238833798%22%2C%22first_id%22%3A%22187599b1de0f7-031709574cc3cf6-26031851-2073600-187599b1de116d%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E8%87%AA%E7%84%B6%E6%90%9C%E7%B4%A2%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTg3NTk5YjFkZTBmNy0wMzE3MDk1NzRjYzNjZjYtMjYwMzE4NTEtMjA3MzYwMC0xODc1OTliMWRlMTE2ZCIsIiRpZGVudGl0eV9sb2dpbl9pZCI6IjM4ODMzNzk4In0%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%24identity_login_id%22%2C%22value%22%3A%2238833798%22%7D%2C%22%24device_id%22%3A%22187599b1de0f7-031709574cc3cf6-26031851-2073600-187599b1de116d%22%7D; tyc-user-info=%7B%22state%22%3A%223%22%2C%22vipManager%22%3A%220%22%2C%22mobile%22%3A%2218033864777%22%2C%22isExpired%22%3A%221%22%7D; tyc-user-info-save-time=1681725774649; auth_token=eyJhbGciOiJIUzUxMiJ9.eyJzdWIiOiIxODAzMzg2NDc3NyIsImlhdCI6MTY4MTcyNTc3NiwiZXhwIjoxNjg0MzE3Nzc2fQ.WhbspdtwHE8oAnxqBStxx7WHnS6uhQsO4DuQk4B2Uciy0Jo1Xx62Jc34QQuzPzwEgyT28DASgZOq1tKIA5g9rg; Hm_lpvt_e92c8d65d92d534b0fc290df538b4758=1681725776',
        'Content-Type': 'application/json',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36'
        }

    def replace_ip(self):
        proxy = ip_proxys.replace_ip()

        self.proxys = {
                'http': proxy,
                'https': proxy
            }


    def req_(self, url):
        prox = self.proxys
        print(prox)
        # self.headers['token'] = token
        res = self._session.get(url=url, headers=self.headers,proxies=prox,  verify=False, timeout=60)
        if res.status_code == requests.codes.ok:
            res_data = res.content.decode()
            return res_data


    def request_(self, url):
        try:
            # time.sleep(random.random() * 10)
            res = self.req_(url)
            if not res:
                res=self.req_(url)
                print('None, waiting 10 m')
                print(res)
                time.sleep(random.random() * 10)
                self.replace_ip()
        except Exception:
            logging.info("proxy disabled！！！ change proxy")
            time.sleep(random.random() * 10)
            self.replace_ip()
            res = self.req_(url)
        # print(res)

        return res


    def write_data(self,file,data):
        with open(file,'a',encoding='utf-8') as w:
            w.write(data+'\n')
            w.close()

    # 建设工程企业list
    def channel_info(self,company):
        self.replace_ip()
        channel_url = f'https://www.tianyancha.com/search?key={urllib.parse.quote(company)}&sessionNo=1681781135.83886633'
        print(channel_url)
        channel=self.request_(channel_url)
        if channel:
            html=etree.HTML(channel)
            href=html.xpath("//div[@class='index_name__qEdWi']/a/@href")
            if href:
                href = href[0]
            print(href)
            time.sleep(random.random()*10)
            detail = self.request_(href)
            if detail:
                html = etree.HTML(detail)
                phone = html.xpath("//span[@class='index_detail-tel__fgpsE']/text()")
                if phone:
                    phone=phone[0]
                email = html.xpath("//span[@class='index_detail-email__B_1Tq']/text()")
                if email:
                    email=email[0]
                address = html.xpath("//span[@class='index_detail-address__ZmaTI']/text()")
                if address:
                    address=address[0]
                return [company,phone, email, address]



def save_xls(datalist):
    book = xlwt.Workbook(encoding="utf-8", style_compression=0)
    sheet = book.add_sheet('河北二级数据', cell_overwrite_ok=True)
    caption = ['company', 'phone', 'email','address' ]
    print()
    for i in range(0, len(caption)):
        sheet.write(0, i, caption[i])
    print(len(datalist))
    # time.sleep(2222)
    for i in range(len(datalist)):
        print(datalist[i])
        if len(datalist[i]) > 4:
            style = xlwt.easyxf('font: bold on')
            sheet.write(i + 1, 0, datalist[i], style)
        else:
            for j in range(len(caption)):
                print(i + 1, j, datalist[i][j])
                sheet.write(i + 1, j, datalist[i][j])  # 写入一行数据


def read_xlsx(file):
    wb = load_workbook(file)
    sheets = wb.worksheets
    sheet1 = wb['Sheet1']
    # 获取第一行所有数据
    rows = sheet1.rows
    columns = sheet1.columns
    lists=[]
    for col in columns:
        col_list=[]
        for row in col:
            if row.value:
                col_list.append(row.value)
            else:
                break
        print(col_list)
        lists.append(col_list)
    print(lists)
    return lists


def read_txt_wirte_xls():#txt 获取数据 存xls
    with open('companyname/河北三级.txt', 'r', encoding='utf-8') as r:
        datalist=r.readlines()
    print(datalist)
    # time.sleep(2222)
    book = xlwt.Workbook(encoding="utf-8", style_compression=0)
    sheet = book.add_sheet('河北三级数据', cell_overwrite_ok=True)
    caption = ['company', 'phone', 'email', 'address']
    print()
    for i in range(0, len(caption)):
        sheet.write(0, i, caption[i])
    print(len(datalist))
    for i in range(len(datalist)):
        print(type(datalist[i]),datalist[i])
        if '[' in datalist[i]:
            for j in range(len(caption)):
                print(i+1,j,eval(datalist[i])[j])
                sheet.write(i+1,j,eval(datalist[i])[j])  # 写入一行数据
        else:
            style = xlwt.easyxf('font: bold on')
            sheet.write(i + 1, 0, datalist[i], style)
    book.save('./河北三级数据.xls')  # 保存

def run(rfile,wfile,errfile):
    spider = get_company()
    datalist = []
    for lists in read_xlsx(rfile):
        for num in range(len(lists)):
            try:
                if num == 0:
                    print(lists[0])
                    data = lists[0]
                else:
                    print(lists[num])
                    companybase = spider.channel_info(lists[num])
                    print(companybase)
                    data = str(companybase)
                spider.write_data(wfile, data)
            except Exception:
                traceback.print_exc()
                spider.write_data(errfile, lists[num])

if __name__ == '__main__':
    # run('companyname/河北二级.xlsx','companyname/河北二级.txt','河北二error.txt')
    read_txt_wirte_xls()


    # namelist=[
# '河北建工集团有限责任公司',
# '唐山开滦建设（集团）有限责任公司',
# '河北承工建设集团有限公司',
# '承德全顺达电子集团有限公司',
# '河北天元亨达智能工程有限责任公司',
# '石家庄市森田电子通信有限公司',
# '石家庄市大业建筑工程有限公司',
# '河北欧意科技集团有限公司']
        # time.sleep(2222)
        # save_xls(datalist)
    # spider=get_company()

    # for cname in namelist:
    #     companybase=spider.channel_info(cname)
    #     spider.write_data('companyname/河北二级1.txt', str(companybase))
    # cell_11 = sheet1.cell(1, 1).value
    # print(cell_11)
    # wb=xlrd.open_workbook('./companyname/')
    # names = wb.sheet_names('河北二级.xlsx')
    # codenum=spider.channel_info('河北佑恩公路工程有限公司')

    # spider.get_company_info(codenum)
